#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
엑셀(.xlsx) 또는 CSV(.csv) 고객 마스터 데이터를 config/companies.json 으로 변환하는 스크립트.
CSV는 파이썬 표준 라이브러리만으로 처리되어 openpyxl 설치 없이도 동작합니다.

사용법
    python scripts/import_companies.py 고객데이터_FY26.csv
    python scripts/import_companies.py 고객데이터_FY26.xlsx --sheet "Sheet1"

지원하는 엑셀 형식
1) 고객 마스터 형식 (권장, 미리 정해진 헤더):
   고객번호 | 고객명 | SUB번호 | SUB고객명 | 팀구분 | 셀코드 | 외근영업 | 내근영업 | 업계
   - (고객번호, 고객명) 을 기준으로 그룹핑합니다. 뉴스 검색은 "고객명"으로만 수행하고,
     같은 그룹에 속한 SUB고객명들은 검색에는 쓰지 않되, 기사 본문/제목에 등장할 때만
     "관련: OO" 형태로 표시하기 위한 참고 정보로 저장합니다.
   - 팀구분/셀코드/외근영업/내근영업/업계는 웹사이트 필터(팀/셀/개인/고객명)에 사용됩니다.
2) 단순 형식 (헤더에 아래 이름 중 하나가 있으면 자동 인식):
   회사명 컬럼: "고객사", "고객사명", "회사명", "name", "company"
   검색어 컬럼(선택): "검색어", "query"
   그룹/담당자 컬럼(선택): "그룹", "담당자", "group"
   위 경우 팀/셀/개인 필터 없이 회사명 중심의 단순 리스트로 저장됩니다.

필요 패키지: CSV는 없음(표준 라이브러리). xlsx는 openpyxl 필요 (pip install openpyxl --break-system-packages)
"""

import sys
import os
import re
import csv
import json
import argparse
from collections import OrderedDict

BASE_DIR = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
OUTPUT_PATH = os.path.join(BASE_DIR, "config", "companies.json")

MASTER_HEADERS = {
    "code": ["고객번호"],
    "name": ["고객명"],
    "sub_code": ["SUB번호"],
    "sub_name": ["SUB고객명"],
    "team": ["팀구분"],
    "cell": ["셀코드"],
    "ext_rep": ["외근영업"],
    "int_rep": ["내근영업"],
    "industry": ["업계"],
}

NAME_HEADERS = {"고객사", "고객사명", "회사명", "name", "company"}
QUERY_HEADERS = {"검색어", "query"}
GROUP_HEADERS = {"그룹", "담당자", "group"}

SUFFIX_TOKENS = ["(주)", "㈜", "(유)", "(재)", "(사)", "(합)", "(주 )", "( 주)"]

# 2026-08-25 추가: TSC6은 금형(mold) 담당팀으로, 이 앱이 다루는 "자동화 시장" 영업지원
# 대상이 아니므로 고객 마스터를 companies.json으로 만들 때부터 아예 제외한다.
# (같은 고객사가 TSC6 외 다른 팀에도 걸려 있으면 그 팀 소속 정보는 그대로 유지되고,
# TSC6 소속 행만 제외된다.)
EXCLUDED_TEAMS = {"TSC6"}


def read_rows(path, sheet_name=None):
    """.csv 또는 .xlsx 파일을 읽어 행(tuple) 리스트로 반환합니다.
    CSV는 표준 라이브러리만으로 처리하고, xlsx는 openpyxl이 필요합니다."""
    ext = os.path.splitext(path)[1].lower()
    if ext == ".csv":
        with open(path, "r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            return [tuple(row) for row in reader]

    try:
        from openpyxl import load_workbook
    except ImportError:
        print("openpyxl 패키지가 필요합니다. 다음 명령으로 설치하세요:\n"
              "  pip install openpyxl", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(path, data_only=True)
    ws = wb[sheet_name] if sheet_name else wb.active
    return list(ws.iter_rows(values_only=True))


def find_col_index(headers, names):
    normalized = [str(h).strip() if h else "" for h in headers]
    for idx, h in enumerate(normalized):
        if h in names:
            return idx
    return None


def clean_query(raw_name):
    """고객명에서 뉴스 검색에 쓸 검색어를 뽑아낸다.

    2026-08-24 수정: 예전에는 "(주)"/"㈜" 등 법인 표기를 빈 문자열로 치환해서 지웠는데,
    그러면 "현대자동차(주)울산공장"처럼 법인 표기가 "브랜드명"과 "사업장명" 사이에 낀
    이름은 "현대자동차울산공장"처럼 공백 없이 그대로 붙어버려서, 실제 어떤 기사 본문에도
    그대로 등장할 수 없는 검색어가 만들어지는 문제가 있었다(대기업 사업장인데도 뉴스가
    한 건도 안 잡히는 원인이었음). 법인 표기를 "공백"으로 치환하고, 남아있는 대괄호/괄호
    문자도 공백으로 치환한 뒤 중복 공백을 정리해서, "현대자동차 울산공장"처럼 실제 기사에
    등장할 수 있는 형태에 가깝게 만든다. (원래 이름 자체에 공백이 전혀 없는 복합어,
    예: "기아자동차화성공장"은 이 방식으로는 못 고치며 company_overrides.json의
    search_queries로 개별 대응해야 한다.)
    """
    if not raw_name:
        return raw_name
    name = str(raw_name)
    for token in SUFFIX_TOKENS:
        name = name.replace(token, " ")
    name = re.sub(r"_ERP$", "", name, flags=re.IGNORECASE)
    name = re.split(r"[-_]", name)[0]
    name = re.sub(r"[\[\]\(\)]", " ", name)
    name = re.sub(r"\s+", " ", name).strip()
    return name


def import_master_format(rows, col_idx):
    groups = OrderedDict()

    def get_val(row, key):
        idx = col_idx.get(key)
        if idx is None or idx >= len(row):
            return None
        v = row[idx]
        return str(v).strip() if v is not None else None

    for row in rows:
        if not row:
            continue

        code = get_val(row, "code")
        name = get_val(row, "name")
        if not name:
            continue
        sub_name = get_val(row, "sub_name")
        team = get_val(row, "team")
        cell = get_val(row, "cell")
        ext_rep = get_val(row, "ext_rep")
        int_rep = get_val(row, "int_rep")
        industry = get_val(row, "industry")

        if team in EXCLUDED_TEAMS:
            # 금형(TSC6) 담당 행은 이 앱의 대상이 아니므로 통째로 건너뛴다.
            # 이 회사의 다른 SUB 행이 TSC6이 아닌 다른 팀이면 그 행은 정상적으로 처리된다.
            continue

        key = (code or "", name)
        if key not in groups:
            groups[key] = {
                "code": code or "",
                "name": name,
                "query": clean_query(name),
                "sub_names": OrderedDict(),
                "team": OrderedDict(),
                "cell": OrderedDict(),
                "ext_rep": OrderedDict(),
                "int_rep": OrderedDict(),
                "industry": OrderedDict(),
            }
        g = groups[key]

        if sub_name and sub_name != name and clean_query(sub_name) != g["query"]:
            g["sub_names"][sub_name] = clean_query(sub_name)
        if team:
            g["team"][team] = True
        if cell:
            g["cell"][cell] = True
        if ext_rep:
            g["ext_rep"][ext_rep] = True
        if int_rep:
            g["int_rep"][int_rep] = True
        if industry:
            g["industry"][industry] = True

    companies = []
    for g in groups.values():
        companies.append({
            "code": g["code"],
            "name": g["name"],
            "query": g["query"],
            "sub_names": [
                {"raw": raw, "clean": clean} for raw, clean in g["sub_names"].items()
            ],
            "team": list(g["team"].keys()),
            "cell": list(g["cell"].keys()),
            "ext_rep": list(g["ext_rep"].keys()),
            "int_rep": list(g["int_rep"].keys()),
            "industry": list(g["industry"].keys()),
        })
    return companies


def import_simple_format(rows, name_idx, query_idx, group_idx):
    companies = []
    seen_names = set()
    for row in rows:
        if row is None or name_idx >= len(row):
            continue
        name = row[name_idx]
        if not name or not str(name).strip():
            continue
        name = str(name).strip()
        if name in seen_names:
            continue
        seen_names.add(name)

        query = name
        if query_idx is not None and query_idx < len(row) and row[query_idx]:
            query = str(row[query_idx]).strip()

        group = ""
        if group_idx is not None and group_idx < len(row) and row[group_idx]:
            group = str(row[group_idx]).strip()

        companies.append({
            "code": "",
            "name": name,
            "query": query,
            "sub_names": [],
            "team": [],
            "cell": [],
            "ext_rep": [],
            "int_rep": [group] if group else [],
            "industry": [],
        })
    return companies


def main():
    parser = argparse.ArgumentParser(description="엑셀(.xlsx) 또는 CSV(.csv) 고객(사) 리스트를 companies.json으로 변환")
    parser.add_argument("xlsx_path", help="고객 목록이 담긴 .xlsx 또는 .csv 파일 경로")
    parser.add_argument("--sheet", default=None, help="시트 이름 (xlsx에만 해당, 생략 시 활성 시트 사용)")
    args = parser.parse_args()

    rows = read_rows(args.xlsx_path, args.sheet)
    if not rows:
        print("파일에 데이터가 없습니다.", file=sys.stderr)
        sys.exit(1)

    header = rows[0]
    data_rows = rows[1:]

    master_idx = {}
    for key, candidates in MASTER_HEADERS.items():
        idx = find_col_index(header, set(candidates))
        if idx is not None:
            master_idx[key] = idx

    if "name" in master_idx and "sub_name" in master_idx:
        print("고객 마스터 형식으로 인식했습니다 (고객명/SUB고객명 등).")
        companies = import_master_format(data_rows, master_idx)
    else:
        print("단순 회사명 목록 형식으로 인식했습니다.")
        name_idx = find_col_index(header, NAME_HEADERS)
        query_idx = find_col_index(header, QUERY_HEADERS)
        group_idx = find_col_index(header, GROUP_HEADERS)
        if name_idx is None:
            name_idx = 0
            data_rows = rows
            print("헤더에서 회사명 컬럼을 찾지 못해 첫 번째 컬럼을 회사명으로 사용합니다.")
        companies = import_simple_format(data_rows, name_idx, query_idx, group_idx)

    if not companies:
        print("변환할 고객사 데이터를 찾지 못했습니다.", file=sys.stderr)
        sys.exit(1)

    output = {
        "_comment": "담당 고객사 목록. scripts/import_companies.py 로 엑셀에서 자동 생성되었습니다.",
        "companies": companies,
    }

    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        json.dump(output, f, ensure_ascii=False, indent=2)

    print(f"완료: {len(companies)}개 고객사(그룹) -> {OUTPUT_PATH}")


if __name__ == "__main__":
    main()
